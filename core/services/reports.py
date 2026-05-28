"""Servicio Reports — generación de Excels (monitoring multi-hoja + export simple).

Port del report_service del DocFlow grande, sin dependencia de notification_service
ni de SharedDrive. La GUI ofrece File dialog para elegir dónde guardar.
"""

import io
import logging
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import DataBarRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ─── Constantes de estilo (idénticas a DocuControl) ───────────────────────────

FILL_HEADER = PatternFill("solid", fgColor="6678AF")
FILL_ROW = PatternFill("solid", fgColor="D4DCF4")
FONT_WHITE = Font(color="FFFFFF", bold=True)
FONT_BLACK = Font(color="000000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_VCE = Alignment(vertical="center")

ESTADO_COLORS = {
    "Rechazado":    "FFA19A",
    "Com. Menores": "FFE5AD",
    "Com. Mayores": "DBB054",
    "Comentado":    "F79646",
    "Enviado":      "B1E1B9",
    "Sin Enviar":   "FFFFAB",
    "Información":  "FFFF46",
    "HOLD":         "FF0909",
    "Aprobado":     "00D25F",
}

RESPONSABLE_COLORS = {
    "SS": "262626", "CCH": "00B0F0", "JM": "2C8A6A", "JV": "006B95",
    "EC": "B31274", "ES": "5A0DA0", "JP": "00458F", "AC": "3F0075",
    "LB": "176DD1", "RM": "228B22", "RP": "1B365D", "EC/SS": "1F7A1F",
}

TAB_COLORS = {
    "ALL DOC.":      "6678AF",
    "ENVIADOS":      "00D25F",
    "DEVOLUCIONES":  "FFA19A",
    "CRÍTICOS":      "DBB054",
    "CRÍTICOS +15d": "FF7F50",
    "SIN ENVIAR":    "FFFF66",
    "STATUS GLOBAL": "B1E1B9",
}

COLUMN_ORDER = [
    "Nº Pedido", "Responsable", "Nº Oferta", "Nº PO", "Cliente", "Material",
    "Fecha Pedido", "Fecha Prevista", "Nº Doc. Cliente", "Nº Doc. EIPSA",
    "Título", "Tipo Doc.", "Info/Review", "Repsonsable", "Días Envío",
    "Crítico", "Estado", "Notas", "Nº Revisión",
    "Fecha Env. Doc.", "Fecha Dev. Doc.", "Días Devolución",
    "Reclamaciones", "Seguimiento", "Historial Rev.",
]

COLS_CENTRAR = {
    "Responsable", "Repsonsable", "Info/Review", "Días Envío", "Crítico",
    "Estado", "Días Devolución", "Nº Revisión",
    "Aprobado", "Com. Mayores", "Com. Menores", "Enviado", "Rechazado",
    "Sin Enviar", "Total", "% Completado",
}

DATE_COLS = {"Fecha Pedido", "Fecha Prevista", "Fecha Env. Doc.", "Fecha Dev. Doc."}


# ═══════════════════════════════════════════════════════════════════════════
#  API pública
# ═══════════════════════════════════════════════════════════════════════════

def generate_monitoring_excel(sections: dict) -> bytes:
    """Excel multi-hoja con secciones (ALL/ENVIADOS/DEVOLUCIONES/...) + STATUS GLOBAL."""
    all_docs = sections.get("all_docs") or (
        sections.get("enviados", [])
        + sections.get("devoluciones", [])
        + sections.get("sin_enviar", [])
        + [
            d for d in sections.get("criticos", [])
            if not any(
                d.get("Nº Doc. EIPSA") == x.get("Nº Doc. EIPSA")
                for x in sections.get("enviados", []) + sections.get("devoluciones", [])
            )
        ]
    )
    enviados = sections.get("enviados", [])
    devoluciones = sections.get("devoluciones", [])
    sin_enviar = sections.get("sin_enviar", [])

    criticos_raw = [
        d for d in sections.get("criticos", [])
        if str(d.get("Estado", "") or "").strip().lower() != "enviado"
    ]
    criticos = [_normalize_pending(d) for d in criticos_raw
                if _dias_dev(d) is None or _dias_dev(d) <= 15]
    criticos_15d = [_normalize_pending(d) for d in criticos_raw
                    if _dias_dev(d) is not None and _dias_dev(d) > 15]

    sin_enviar = [_normalize_pending(d) for d in sin_enviar]
    devoluciones = [_add_notas(d) for d in devoluciones]

    # Normalizar Estado vacío → "Sin Enviar" en TODAS las hojas restantes.
    # No tocamos Días Devolución porque enviados/devoluciones SÍ tienen valor
    # calculado y CRÍTICOS / SIN ENVIAR ya pasaron por _normalize_pending.
    all_docs = [_normalize_estado(d) for d in all_docs]
    enviados = [_normalize_estado(d) for d in enviados]
    devoluciones = [_normalize_estado(d) for d in devoluciones]

    wb = Workbook()
    wb.remove(wb.active)

    sheet_order = [
        ("ALL DOC.",      all_docs),
        ("ENVIADOS",      enviados),
        ("DEVOLUCIONES",  devoluciones),
        ("CRÍTICOS",      criticos),
        ("CRÍTICOS +15d", criticos_15d),
        ("SIN ENVIAR",    sin_enviar),
    ]

    for sheet_name, data in sheet_order:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = TAB_COLORS[sheet_name]

        cols = _get_cols(data)
        _write_header(ws, cols)
        _write_rows(ws, data, cols)
        _apply_cell_styles(ws, cols)

        if sheet_name == "DEVOLUCIONES" and "Días Devolución" in cols:
            col_idx = cols.index("Días Devolución") + 1
            col_letter = get_column_letter(col_idx)
            max_col_letter = get_column_letter(len(cols))
            if ws.max_row >= 2:
                diff_row = DifferentialStyle(
                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                )
                formula = f"=${col_letter}2>15"
                rule_row = Rule(type="expression", formula=[formula], dxf=diff_row, stopIfTrue=False)
                ws.conditional_formatting.add(f"A2:{max_col_letter}{ws.max_row}", rule_row)

        _finalize_sheet(ws, cols)

    _build_status_global_sheet(wb, all_docs)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_export_excel(documents: list[dict], filters: dict | None = None) -> bytes:
    """Excel simple: 1 hoja con datos + 1 hoja con resumen."""
    df = pd.DataFrame(documents)
    if filters:
        for key, value in filters.items():
            if value and key in df.columns:
                df = df[df[key].astype(str).str.contains(str(value), case=False, na=False)]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Documentos")
        summary = _build_summary(df)
        pd.DataFrame(summary).to_excel(writer, index=False, sheet_name="Resumen")
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
#  STATUS GLOBAL sheet
# ═══════════════════════════════════════════════════════════════════════════

def _build_status_global_sheet(wb: Workbook, all_docs: list) -> None:
    if not all_docs:
        return

    df = pd.DataFrame(all_docs)
    if "Nº Pedido" not in df.columns or "Estado" not in df.columns:
        return

    df["Estado"] = df["Estado"].fillna("Sin Enviar").astype(str)

    status_global = (
        df.groupby(["Nº Pedido", "Estado"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for col in ["Aprobado", "Com. Mayores", "Com. Menores", "Enviado", "Rechazado", "Sin Enviar"]:
        if col not in status_global.columns:
            status_global[col] = 0
    status_global["Total"] = status_global.iloc[:, 1:].sum(axis=1)
    status_global["% Completado"] = (
        status_global.get("Aprobado", 0) / status_global["Total"] * 100
    ).fillna(0).round(2)

    status_global = status_global[status_global["% Completado"] != 100].copy()
    status_global.sort_values("Nº Pedido", ascending=False, inplace=True)

    sg_cols = ["Nº Pedido", "Aprobado", "Com. Mayores", "Com. Menores",
               "Enviado", "Rechazado", "Sin Enviar", "Total", "% Completado"]
    sg_cols = [c for c in sg_cols if c in status_global.columns]

    ws = wb.create_sheet(title="STATUS GLOBAL")
    ws.sheet_properties.tabColor = TAB_COLORS["STATUS GLOBAL"]

    _write_header(ws, sg_cols)

    fill_sg = PatternFill("solid", fgColor="D4DCF4")
    for _, row_data in status_global.iterrows():
        ws.append([row_data.get(c, 0) for c in sg_cols])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_sg
            cell.border = THIN_BORDER
            cell.font = FONT_BLACK
            col_name = sg_cols[cell.column - 1] if cell.column <= len(sg_cols) else ""
            cell.alignment = ALIGN_CENTER if col_name in COLS_CENTRAR else ALIGN_VCE
            if col_name == "% Completado" and cell.value == 100:
                cell.font = Font(color="FF0000", bold=True)

    pct_idx = sg_cols.index("% Completado") + 1 if "% Completado" in sg_cols else None
    if pct_idx and ws.max_row >= 2:
        pct_letter = get_column_letter(pct_idx)
        ws.conditional_formatting.add(
            f"{pct_letter}2:{pct_letter}{ws.max_row}",
            DataBarRule(
                start_type="percentile", start_value=0,
                end_type="percentile", end_value=100,
                color="4472C4", showValue="1",
            ),
        )

    _finalize_sheet(ws, sg_cols)

    if ws.max_row >= 3:
        estado_cols_chart = ["Aprobado", "Com. Mayores", "Com. Menores",
                             "Enviado", "Rechazado", "Sin Enviar"]
        chart_col_indices = [sg_cols.index(c) + 1 for c in estado_cols_chart if c in sg_cols]
        if chart_col_indices:
            min_col_c = min(chart_col_indices)
            max_col_c = max(chart_col_indices)

            chart = BarChart()
            chart.type = "col"
            chart.title = "Estado por Pedido"
            chart.style = 12
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.y_axis.title = "Nº Documentos"
            chart.x_axis.title = "Nº Pedido"

            data_ref = Reference(ws, min_col=min_col_c, max_col=max_col_c,
                                 min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            colores_chart = ["00B350", "C59B3F", "FFCF7F", "5566A0", "FF8273", "FFEF7F"]
            for idx, serie in enumerate(chart.series):
                if idx < len(colores_chart):
                    serie.graphicalProperties = GraphicalProperties(
                        solidFill=colores_chart[idx],
                    )

            chart.height = 16
            chart.width = 29
            ws.add_chart(chart, "J3")


# ═══════════════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _dias_dev(doc: dict) -> int | None:
    v = doc.get("Días Devolución")
    try:
        return int(float(v)) if v not in (None, "", "nan") else None
    except (ValueError, TypeError):
        return None


def _normalize_estado(doc: dict) -> dict:
    """Rellena Estado='Sin Enviar' si está vacío. No toca el resto de campos.

    Se aplica a todas las hojas (ALL DOC., ENVIADOS, DEVOLUCIONES, etc.) para
    que ningún documento aparezca con la celda Estado vacía en el Excel.
    """
    estado_raw = str(doc.get("Estado") or "").strip()
    if not estado_raw or estado_raw.lower() == "nan":
        d = dict(doc)
        d["Estado"] = "Sin Enviar"
        return d
    return doc


def _normalize_pending(doc: dict) -> dict:
    """Para docs Sin Enviar: rellena Estado='Sin Enviar' y Días Devolución=0.

    Aplica a las hojas CRÍTICOS, CRÍTICOS +15d y SIN ENVIAR para evitar celdas
    vacías cuando un doc todavía no se ha enviado al cliente.
    """
    d = _normalize_estado(doc)
    if d is doc:  # _normalize_estado no clonó (estado ya tenía valor)
        d = dict(doc)
    dd = d.get("Días Devolución", "")
    if dd in (None, "") or str(dd).strip().lower() == "nan":
        d["Días Devolución"] = 0
    return d


def _parse_any_date(val) -> datetime | None:
    """Intenta parsear una fecha en cualquier formato típico del ERP."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    head = s.split("T")[0] if "T" in s else s.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(head, fmt)
        except ValueError:
            continue
    return None


def _add_notas(doc: dict) -> dict:
    """Añade columna 'Notas' a una devolución: 'Enviar antes del DD/MM/YYYY'."""
    d = dict(doc)
    dt = _parse_any_date(d.get("Fecha Env. Doc.") or d.get("Fecha"))
    if dt:
        limite = dt + timedelta(days=15)
        d["Notas"] = f"Enviar antes del {limite.strftime('%d/%m/%Y')}"
    else:
        d["Notas"] = ""
    return d


def _get_cols(data: list) -> list[str]:
    if not data:
        return list(COLUMN_ORDER)
    available = set(data[0].keys())
    ordered = [c for c in COLUMN_ORDER if c in available]
    extras = [c for c in data[0].keys() if c not in ordered]
    return ordered + extras


def _fmt_value(val, col_name: str):
    if val is None or val == "":
        return ""
    if col_name in DATE_COLS:
        if isinstance(val, datetime):
            return val
        try:
            clean = str(val).split("T")[0]
            return datetime.strptime(clean, "%Y-%m-%d")
        except Exception:
            return str(val)
    if col_name in {"Días Envío", "Días Devolución", "Nº Revisión"}:
        try:
            n = int(float(val))
            return n if str(val) not in ("", "nan") else ""
        except (ValueError, TypeError):
            return str(val)
    return str(val) if not isinstance(val, (int, float, bool)) else val


def _write_header(ws, cols: list) -> None:
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER


def _write_rows(ws, data: list, cols: list) -> None:
    for doc in data:
        row_vals = [_fmt_value(doc.get(c), c) for c in cols]
        ws.append(row_vals)
        for ci, col_name in enumerate(cols, start=1):
            if col_name in DATE_COLS:
                cell = ws.cell(ws.max_row, ci)
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD/MM/YYYY"


def _apply_cell_styles(ws, cols: list) -> None:
    idx = {name: i for i, name in enumerate(cols)}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = FILL_ROW
            cell.border = THIN_BORDER
            cell.font = FONT_BLACK
            col_name = cols[cell.column - 1] if cell.column <= len(cols) else ""
            cell.alignment = ALIGN_CENTER if col_name in COLS_CENTRAR else ALIGN_VCE

        if "Estado" in idx:
            estado_cell = row[idx["Estado"]]
            estado_val = str(estado_cell.value or "")
            if estado_val in ESTADO_COLORS:
                estado_cell.fill = PatternFill("solid", fgColor=ESTADO_COLORS[estado_val])
            if estado_val == "HOLD":
                estado_cell.font = Font(color="FF0000", bold=True)
            elif estado_val == "Aprobado":
                estado_cell.font = Font(bold=True)

        for resp_col in ("Responsable", "Repsonsable"):
            if resp_col in idx:
                resp_cell = row[idx[resp_col]]
                resp_val = str(resp_cell.value or "")
                if resp_val in RESPONSABLE_COLORS:
                    resp_cell.font = Font(color=RESPONSABLE_COLORS[resp_val], bold=True)

        if "Crítico" in idx:
            crit_cell = row[idx["Crítico"]]
            if str(crit_cell.value or "").strip().lower() in ("sí", "si"):
                crit_cell.font = Font(color="FF0000", bold=True)

        if "Info/Review" in idx:
            ir_cell = row[idx["Info/Review"]]
            if ir_cell.value == "R":
                ir_cell.font = Font(color="FF0000", bold=True)
            elif ir_cell.value == "I":
                ir_cell.font = Font(color="4D4D4D", bold=True)

        if "Días Devolución" in idx:
            dd_cell = row[idx["Días Devolución"]]
            try:
                if dd_cell.value is not None and int(float(dd_cell.value)) > 15:
                    dd_cell.font = Font(color="FF0000", bold=True)
            except (ValueError, TypeError):
                pass


def _finalize_sheet(ws, cols: list) -> None:
    ws.freeze_panes = "B2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for col_idx, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _build_summary(df: pd.DataFrame) -> list[dict]:
    rows = [{"Metrica": "Total documentos", "Valor": len(df)}]
    for col in df.columns:
        if col.lower() in ("estado", "status"):
            for status, count in df[col].value_counts().items():
                rows.append({"Metrica": f"Estado: {status}", "Valor": int(count)})
    return rows
