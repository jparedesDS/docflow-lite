"""Apertura de pedidos — localiza el pedido existente y añade 00 DOCUMENTACIÓN + Planning + VDDL.

Flujo (basado en cómo trabajamos en EIPSA):

1. La carpeta del pedido **ya existe** en `M:\\base de datos de pedidos\\Año <YYYY>\\
   <YYYY> Pedidos\\` (la crea otro proceso). Ejemplos reales:
     - `P-26-003 - BP OIL REFINERIA - ELEMENTO TEMPERATURA`
     - `P-26-050-S00 - SABIC - NIVEL`
   El servicio localiza esa carpeta por prefijo (`P-XX-XXX`).

2. Identifica la subcarpeta de trabajo técnico (`2-Tecnico`, `02 tecnico`,
   `02-Tecnico`, etc.). Si no existe se crea como `2-Tecnico`.

3. Crea `00 DOCUMENTACIÓN` dentro de la técnica y copia el contenido íntegro de
   `M:\\Comunes\\JOSE\\00 DOCUMENTACIÓN\\00 DOCUMENTACIÓN`.

4. Genera el Planning (copia + rellena hoja PLAN ING con fechas y S.REF/N.REF).

5. Genera un VDDL básico con openpyxl.

Sin BBDD, sin red automática — todo en local + red mapeada (M:).
"""

from __future__ import annotations

import logging
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# Rutas por defecto (configurables al llamar a create_order)
# ════════════════════════════════════════════════════════════════════════════

DEFAULT_BASE_DIR = Path(r"M:\base de datos de pedidos")
DEFAULT_TEMPLATE_DIR = Path(r"M:\Comunes\JOSE\00 DOCUMENTACIÓN\00 DOCUMENTACIÓN")
DEFAULT_PLANNING_TEMPLATE = (
    DEFAULT_TEMPLATE_DIR / "03 PLANTILLA PLANNING - CON MESES.xlsm"
)
DEFAULT_ERP_TEMPLATE = (
    DEFAULT_TEMPLATE_DIR / "01 IMPORTAR DOCUMENTOS - ERP - sin enviar.xlsx"
)


# ════════════════════════════════════════════════════════════════════════════
# Catálogo de subcarpetas "env. *" dentro de 2-Tecnico
# ════════════════════════════════════════════════════════════════════════════
#
# Single source of truth: la misma lista gobierna
#   1) qué subcarpetas se crean en 2-Tecnico
#   2) qué filas se generan en la VDDL automática (1 fila por subcarpeta)
#
# Para cada entrada:
#   folder : nombre EXACTO de la carpeta a crear en disco
#   es     : Título de la fila VDDL cuando el cliente es español
#   en     : Título de la fila VDDL cuando el cliente es inglés
#   code   : "Tipo Documento" del ERP — folder shortname sin "env. "
#
# NOTA: env. INDEX QUALITY DOSSIER no tenía mapping ES en la spec del usuario;
# usamos un fallback razonable derivado del nombre.

# Opciones EXACTAS del desplegable "Tipo Documento" en la plantilla ERP
# (extraídas con data_validation de openpyxl). Debe coincidir LITERAL —
# si no, Excel marca la celda como inválida.
ERP_TIPO_DOC_OPTIONS: tuple[str, ...] = (
    "Cálculo y plano", "Cálculos", "Catálogo", "Certificados", "Datos técnicos",
    "Dossier", "Índice", "Informe", "Instrucciones", "Listado", "Manual",
    "Nameplate", "Otros", "Packing", "Planos", "PMI", "PPI", "Procedimientos",
    "Programa", "Repuestos", "Soldadura", "VDDL",
)


def _fold(s: str) -> str:
    """Normaliza string a minúsculas sin acentos ni espacios extra, para comparar."""
    import unicodedata
    s = "".join(
        c for c in unicodedata.normalize("NFD", str(s or ""))
        if not unicodedata.combining(c)
    )
    return s.lower().strip().replace("\xa0", " ")


def _canonical_tipo_doc(raw: str) -> str:
    """Devuelve el valor EXACTO del desplegable que más se parece a `raw`.

    Tolera typos del usuario: case (Tecnicos vs técnicos), acentos, nbsp, espacios.
    Si no encuentra match devuelve el valor original (la celda quedará inválida
    en Excel y validate_catalog avisará).
    """
    if raw is None:
        return ""
    s = str(raw).strip().replace("\xa0", " ")
    if s in ERP_TIPO_DOC_OPTIONS:
        return s
    target = _fold(s)
    for opt in ERP_TIPO_DOC_OPTIONS:
        if _fold(opt) == target:
            return opt
    return s

# Catálogo de subcarpetas con triple mapping:
#   tipo_doc   → valor LITERAL del desplegable ERP (columna E) — pasa por _canonical_tipo_doc
#   eipsa_code → identificador EIPSA "XXX-NNNN" (parte central del Nº Doc EIPSA columna A)
#   es / en    → Título de la columna D según idioma del cliente
#
# El Nº Doc EIPSA final se construye como  "{YY-NNN}-{eipsa_code}"
# Ejemplo: pedido P-26-050 + env. VDDL (LIS-0001) → "26-050-LIS-0001"
#
# Fuente: codes.xlsx proporcionado por el Document Controller (30 subcarpetas).
SUBFOLDER_CATALOG: list[dict] = [
    {"folder": "env. Cálculos y Planos",
     "es": "Cálculos y Planos Dimensionales",   "en": "Calculations and Dimensional Drawings",
     "tipo_doc": _canonical_tipo_doc("Cálculo y plano"),     "eipsa_code": "ESP-0003",
     "critico": "Sí"},
    {"folder": "env. Cálculos",
     "es": "Cálculos",                          "en": "Calculations",
     "tipo_doc": _canonical_tipo_doc("Cálculos"),            "eipsa_code": "CAL-0001",
     "critico": "Sí"},
    {"folder": "env. Planos",
     "es": "Planos Dimensionales",              "en": "Dimensional Drawings",
     "tipo_doc": _canonical_tipo_doc("Planos"),              "eipsa_code": "PLG-0005",
     "critico": "Sí"},
    {"folder": "env. Certificados Calibración",
     "es": "Certificados de calibración",       "en": "Calibration Certificates",
     "tipo_doc": _canonical_tipo_doc("Certificados"),        "eipsa_code": "CER-0006",
     "critico": "No"},
    {"folder": "env. Certificados Eléctricos",
     "es": "Certificados eléctricos",           "en": "Electrical Certificates",
     "tipo_doc": _canonical_tipo_doc("Certificados"),        "eipsa_code": "CER-0001",
     "critico": "No"},
    {"folder": "env. Custom Classification",
     "es": "Custom Classification",             "en": "Custom Classification",
     "tipo_doc": _canonical_tipo_doc("Otros"),               "eipsa_code": "LIS-0022",
     "critico": "No"},
    {"folder": "env. FINAL QUALITY DOSSIER",
     "es": "Dosier Final",                      "en": "Final Quality Dossier",
     "tipo_doc": _canonical_tipo_doc("Dossier"),             "eipsa_code": "DOS-0002",
     "critico": "No"},
    {"folder": "env. INDEX QUALITY DOSSIER",
     "es": "Índice del Dosier Final",           "en": "Final Quality Dossier Index",
     "tipo_doc": _canonical_tipo_doc("Índice"),              "eipsa_code": "DOS-0003",
     "critico": "Sí"},
    {"folder": "env. ITP",
     "es": "Plan de Puntos de Inspección",      "en": "Inspection and Test Plan (ITP)",
     "tipo_doc": _canonical_tipo_doc("PPI"),                 "eipsa_code": "PLN-0001",
     "critico": "Sí"},
    {"folder": "env. List of Hazardous",
     "es": "List of Hazardous",                 "en": "List of Hazardous",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "LIS-0029",
     "critico": "No"},
    {"folder": "env. Manual",
     "es": "MANUAL",                            "en": "Manual",
     "tipo_doc": _canonical_tipo_doc("Manual"),              "eipsa_code": "MAN-0001",
     "critico": "Sí"},
    {"folder": "env. Packing List",
     "es": "Packing List",                      "en": "Packing List",
     "tipo_doc": _canonical_tipo_doc("Packing"),             "eipsa_code": "PRC-0012",
     "critico": "No"},
    {"folder": "env. PH (Prueba Hidrostática)",
     "es": "Prueba Hidrostática",               "en": "Hydrostatic Test",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0010",
     "critico": "No"},
    {"folder": "env. PMI",
     "es": "PMI",                               "en": "Positive Material Identification (PMI)",
     "tipo_doc": _canonical_tipo_doc("PMI"),                 "eipsa_code": "PRC-0008",
     "critico": "No"},
    {"folder": "env. Preliminary Cargo List",
     "es": "Preliminary Cargo List",            "en": "Preliminary Cargo List",
     "tipo_doc": _canonical_tipo_doc("Packing"),             "eipsa_code": "LIS-0024",
     "critico": "No"},
    {"folder": "env. Preservation and Storage",
     "es": "Preservation and Storage",          "en": "Preservation and Storage",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0009",
     "critico": "No"},
    {"folder": "env. Proc. Soldadura",
     "es": "Procedimientos de Soldadura",       "en": "Welding Procedures",
     "tipo_doc": _canonical_tipo_doc("Soldadura"),           "eipsa_code": "DOS-0001",
     "critico": "Sí"},
    {"folder": "env. SPARE PARTS PRECOMMISSIONING",
     "es": "Repuestos para puesta en marcha",   "en": "Spare Parts (Precommissioning)",
     "tipo_doc": _canonical_tipo_doc("Repuestos"),           "eipsa_code": "LIS-0016",
     "critico": "No"},
    {"folder": "env. VDDL",
     "es": "Lista de documentos",               "en": "Vendor Document Deliverables List (VDDL)",
     "tipo_doc": _canonical_tipo_doc("VDDL"),                "eipsa_code": "LIS-0001",
     "critico": "Sí"},
    {"folder": "env. Progress Report",
     "es": "Reporte de Progreso",               "en": "Progress Report",
     "tipo_doc": _canonical_tipo_doc("Informe"),             "eipsa_code": "PRG-0001",
     "critico": "No"},
    {"folder": "env. Proc. TEST",
     "es": "Procedimientos de pruebas y test",  "en": "Test Procedures",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0001",
     "critico": "No"},
    {"folder": "env. Manufacturing Program",
     "es": "Planning de Fabricación",           "en": "Manufacturing Program",
     "tipo_doc": _canonical_tipo_doc("Programa"),            "eipsa_code": "PRG-0003",
     "critico": "Sí"},
    # ─── Subcarpetas añadidas (codes.xlsx) ───────────────────────────────
    {"folder": "env. Certificados Prueba y Materiales",
     "es": "Certificados de pruebas y materiales",
     "en": "Test and Material Certificates",
     "tipo_doc": _canonical_tipo_doc("Certificados"),        "eipsa_code": "CER-0004",
     "critico": "No"},
    {"folder": "env. HOJAS DE DATOS",
     "es": "Hojas de datos",                    "en": "Datasheets",
     "tipo_doc": _canonical_tipo_doc("Datos técnicos"),      "eipsa_code": "HD-0001",
     "critico": "Sí"},
    {"folder": "env. NDE",
     "es": "Procedimientos NDE",                "en": "NDE Procedures",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0006",
     "critico": "No"},
    {"folder": "env. INDEX VENDOR DATA BOOK",
     "es": "Índice del Vendor Data Book",       "en": "Vendor Data Book Index",
     "tipo_doc": _canonical_tipo_doc("Índice"),              "eipsa_code": "VDB-0001",
     "critico": "No"},
    {"folder": "env. VENDOR DATA BOOK",
     "es": "Vendor Data Book",                  "en": "Vendor Data Book",
     "tipo_doc": _canonical_tipo_doc("Dossier"),             "eipsa_code": "VDB-0002",
     "critico": "No"},
    {"folder": "env. SPARE PARTS 2YEARS",
     "es": "Repuestos para 2 años",             "en": "Spare Parts (2 Years)",
     "tipo_doc": _canonical_tipo_doc("Repuestos"),           "eipsa_code": "LIS-0017",
     "critico": "No"},
    {"folder": "env. PINTURA",
     "es": "Procedimiento de pintura",          "en": "Painting Procedure",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0016",
     "critico": "Sí"},
    {"folder": "env. Proc. Calibración",
     "es": "Procedimientos de calibración",     "en": "Calibration Procedures",
     "tipo_doc": _canonical_tipo_doc("Procedimientos"),      "eipsa_code": "PRC-0025",
     "critico": "No"},
]

# Opciones válidas para la columna F del ERP
ERP_CRITICO_OPTIONS: tuple[str, ...] = ("No", "Sí")

# Acceso rápido por nombre de carpeta
SUBFOLDER_INDEX: dict[str, dict] = {e["folder"]: e for e in SUBFOLDER_CATALOG}

# Lista de nombres tal cual están en disco (preserva orden de catálogo)
ALL_SUBFOLDERS: list[str] = [e["folder"] for e in SUBFOLDER_CATALOG]


def subfolder_title(entry: dict, lang: str = "es") -> str:
    """Devuelve el Título VDDL para la subcarpeta en el idioma indicado."""
    lang = (lang or "es").lower()
    return entry.get("en" if lang == "en" else "es") or entry["folder"]


def validate_catalog() -> list[str]:
    """Comprueba que el catálogo está bien formado contra los desplegables ERP."""
    errors: list[str] = []
    valid_tipo = set(ERP_TIPO_DOC_OPTIONS)
    valid_critico = set(ERP_CRITICO_OPTIONS)
    for e in SUBFOLDER_CATALOG:
        if e["tipo_doc"] not in valid_tipo:
            errors.append(
                f"{e['folder']}: tipo_doc={e['tipo_doc']!r} no está en el desplegable ERP."
            )
        critico = e.get("critico")
        if not critico:
            errors.append(f"{e['folder']}: critico vacío.")
        elif critico not in valid_critico:
            errors.append(
                f"{e['folder']}: critico={critico!r} no es 'Sí'/'No'."
            )
        if not e.get("eipsa_code"):
            errors.append(f"{e['folder']}: eipsa_code vacío.")
    return errors

# Regex para validar P-XX-XXX (formato carpetas) y P-XX/XXX (formato N.REF)
_PEDIDO_FOLDER_RE = re.compile(r"^P-(\d{2})-(\d{3})$", re.IGNORECASE)
_PEDIDO_NREF_RE = re.compile(r"^P-(\d{2})/(\d{3})(?:-S\d{2})?$", re.IGNORECASE)

# Tolerante: acepta P-26-003, P-26/003, P-26-003-S00, P-26/003-S00…
_PEDIDO_FLEX_RE = re.compile(
    r"^\s*P-?(\d{2})[\-/](\d{3})(?:\s*[\-/]?\s*(S\d{2}))?\s*$",
    re.IGNORECASE,
)

# Match del comienzo del nombre de una carpeta de pedido — ignora si tiene -S00 o no
# Ejemplos:
#   'P-26-003 - BP OIL REFINERIA - ELEMENTO TEMPERATURA'
#   'P-26-050-S00 - SABIC - NIVEL'
_PEDIDO_DIR_RE = re.compile(
    r"^P-(\d{2})-(\d{3})(?:-S(\d{2}))?\s*[\-–—]\s*(.+)$",
    re.IGNORECASE,
)

# Subcarpeta técnica: '2-Tecnico', '02 tecnico', '02-Tecnico', '02 Técnico', '2 tecnico'…
_TECNICO_RE = re.compile(r"^0?2[\s\-_]+t[eé]cnico\s*$", re.IGNORECASE)


def parse_pedido(raw: str) -> tuple[str, str | None]:
    """Normaliza un pedido a (folder_id, suffix).

    folder_id en formato canónico 'P-XX-XXX'. suffix puede ser None.
    Acepta variantes con barra, con sufijo Sxx, con/sin espacios.
    """
    m = _PEDIDO_FLEX_RE.match(raw or "")
    if not m:
        raise ValueError(
            f"Pedido inválido: {raw!r}. Esperado formato P-XX-XXX (p.ej. P-26-003)."
        )
    yy, nnn, suf = m.group(1), m.group(2), m.group(3)
    folder_id = f"P-{yy}-{nnn}"
    return folder_id, (suf.upper() if suf else None)


def parse_folder_meta(folder_name: str) -> dict | None:
    """Extrae metadatos del nombre de carpeta de un pedido.

    Devuelve dict con `folder_id` (P-XX-XXX), `suffix` (Sxx o None),
    `cliente` y `material` — o None si el nombre no encaja con el patrón.
    """
    m = _PEDIDO_DIR_RE.match(folder_name)
    if not m:
        return None
    yy, nnn, suf, tail = m.group(1), m.group(2), m.group(3), m.group(4).strip()
    # tail = 'CLIENTE - MATERIAL' (puede haber varios guiones; partimos por el primero ' - ')
    parts = re.split(r"\s+[\-–—]\s+", tail, maxsplit=1)
    cliente = parts[0].strip() if parts else ""
    material = parts[1].strip() if len(parts) > 1 else ""
    return {
        "folder_id": f"P-{yy}-{nnn}",
        "suffix": (f"S{suf}" if suf else None),
        "cliente": cliente,
        "material": material,
    }


def find_existing_pedido_dir(
    pedido: str,
    año: int,
    base_dir: Path = DEFAULT_BASE_DIR,
) -> Path | None:
    """Busca una carpeta de pedido existente bajo `Año YYYY\\YYYY Pedidos\\`.

    Empareja por prefijo `P-XX-XXX` ignorando si la carpeta lleva `-S00` o no.
    Si hay varias coincidencias devuelve la primera lexicográficamente.
    """
    folder_id, _ = parse_pedido(pedido)
    año_dir = Path(base_dir) / f"Año {año}" / f"{año} Pedidos"
    if not año_dir.exists():
        return None

    candidates = []
    for entry in año_dir.iterdir():
        if not entry.is_dir():
            continue
        meta = parse_folder_meta(entry.name)
        if meta and meta["folder_id"].upper() == folder_id.upper():
            candidates.append(entry)

    if not candidates:
        return None
    candidates.sort(key=lambda p: p.name)
    return candidates[0]


def find_tecnico_dir(pedido_dir: Path) -> Path | None:
    """Devuelve la subcarpeta técnica del pedido, o None si no existe."""
    if not pedido_dir.exists():
        return None
    for entry in pedido_dir.iterdir():
        if entry.is_dir() and _TECNICO_RE.match(entry.name):
            return entry
    return None


# ════════════════════════════════════════════════════════════════════════════
# Modelos
# ════════════════════════════════════════════════════════════════════════════


@dataclass
class OrderSpec:
    """Especificación de un pedido a procesar.

    `cliente` y `material` son opcionales: si se omiten y la carpeta del
    pedido ya existe en disco, se auto-extraen del nombre de la carpeta.
    """

    pedido: str                 # 'P-26-003'
    fecha_entrada: datetime     # PO sup0 date
    fecha_prevista: datetime    # Final delivery
    cliente: str = ""           # 'BP OIL REFINERIA' (opcional)
    material: str = ""          # 'ELEMENTO TEMPERATURA' (opcional)
    sref: str = ""              # S.REF cliente (PO number externo)
    año: int | None = None      # Por defecto: año de fecha_entrada
    suffix: str = "S00"         # 'S00' default; revisión inicial

    def __post_init__(self) -> None:
        self.cliente = (self.cliente or "").strip()
        self.material = (self.material or "").strip()
        self.sref = (self.sref or "").strip()
        self.suffix = (self.suffix or "").strip().upper() or "S00"

        # Normaliza el pedido aceptando variantes (P-26-003, P-26/003, P-26-003-S00...).
        folder_id, parsed_suffix = parse_pedido(self.pedido)
        self.pedido = folder_id
        # Si el usuario ya incluyó sufijo en el pedido, prevalece sobre el campo separado
        if parsed_suffix:
            self.suffix = parsed_suffix
        if not isinstance(self.fecha_entrada, datetime):
            raise ValueError("fecha_entrada debe ser datetime.")
        if not isinstance(self.fecha_prevista, datetime):
            raise ValueError("fecha_prevista debe ser datetime.")
        if self.fecha_prevista <= self.fecha_entrada:
            raise ValueError("fecha_prevista debe ser posterior a fecha_entrada.")
        if self.año is None:
            self.año = self.fecha_entrada.year

    # ── Helpers derivados ──────────────────────────────────────────────────

    def adopt_from_folder(self, folder_name: str) -> None:
        """Rellena cliente/material/suffix vacíos a partir del nombre de carpeta."""
        meta = parse_folder_meta(folder_name)
        if not meta:
            return
        if not self.cliente:
            self.cliente = meta["cliente"]
        if not self.material:
            self.material = meta["material"]
        if meta["suffix"] and self.suffix == "S00":
            self.suffix = meta["suffix"]

    @property
    def folder_name(self) -> str:
        """'P-26-003 - BP OIL REFINERIA - ELEMENTO TEMPERATURA' (formato canónico para crear nuevas)."""
        return f"{self.pedido} - {self.cliente} - {self.material}"

    @property
    def n_ref(self) -> str:
        """'P-26/003-S00' (formato N.REF interno EIPSA)."""
        m = _PEDIDO_FOLDER_RE.match(self.pedido)
        if not m:
            return self.pedido
        return f"P-{m.group(1)}/{m.group(2)}-{self.suffix}"

    @property
    def n_ref_short(self) -> str:
        """'P-26/021' sin sufijo S00 — para la celda B3 del Planning."""
        m = _PEDIDO_FOLDER_RE.match(self.pedido)
        if not m:
            return self.pedido
        return f"P-{m.group(1)}/{m.group(2)}"

    @property
    def eipsa_prefix(self) -> str:
        """'26-050' — prefijo del Nº Documento EIPSA (sin la 'P-')."""
        m = _PEDIDO_FOLDER_RE.match(self.pedido)
        if not m:
            return self.pedido.lstrip("P-")
        return f"{m.group(1)}-{m.group(2)}"


@dataclass
class OrderResult:
    """Resultado de create_order — rutas creadas + warnings."""

    pedido_dir: Path
    tecnico_dir: Path
    documentacion_dir: Path
    planning_file: Path | None = None
    vddl_file: Path | None = None
    vddl_lang: str = "es"
    subfolders_created: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    template_copied: bool = False


# ════════════════════════════════════════════════════════════════════════════
# Operaciones de filesystem
# ════════════════════════════════════════════════════════════════════════════


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def create_folder_structure(
    spec: OrderSpec,
    base_dir: Path = DEFAULT_BASE_DIR,
    *,
    create_if_missing: bool = True,
) -> tuple[Path, Path, Path]:
    """Localiza la carpeta del pedido y prepara `2-Tecnico\\00 DOCUMENTACIÓN`.

    Estrategia:
      1. Busca la carpeta del pedido por prefijo `P-XX-XXX` (ignora `-S00`).
      2. Si la encuentra, busca la subcarpeta técnica (`2-Tecnico`, `02 tecnico`...).
         Si no existe, crea `2-Tecnico` dentro.
      3. Crea `00 DOCUMENTACIÓN` dentro de la técnica.
      4. Si no encuentra la carpeta del pedido y `create_if_missing=True`,
         la crea con el formato canónico (requiere cliente y material).

    Devuelve (pedido_dir, tecnico_dir, documentacion_dir).
    """
    if not base_dir.exists():
        raise FileNotFoundError(
            f"La carpeta base de pedidos no existe: {base_dir}\n"
            "Comprueba que la unidad de red está conectada."
        )

    año_dir = base_dir / f"Año {spec.año}" / f"{spec.año} Pedidos"

    # 1) Buscar la carpeta del pedido (acepta `-S00` o sin sufijo)
    pedido_dir = find_existing_pedido_dir(spec.pedido, spec.año, base_dir=base_dir)

    if pedido_dir is not None:
        logger.info("Pedido localizado: %s", pedido_dir)
        # Adoptar cliente/material desde la carpeta si no se proveyeron
        spec.adopt_from_folder(pedido_dir.name)
    else:
        # 2) No existe → crearla
        if not create_if_missing:
            raise FileNotFoundError(
                f"No se encontró carpeta de pedido para {spec.pedido} en {año_dir}.\n"
                f"Verifica que el pedido esté creado o activa la creación automática."
            )
        if not spec.cliente or not spec.material:
            raise ValueError(
                f"No se encontró carpeta para {spec.pedido} y faltan Cliente/Material "
                "para crear una nueva."
            )
        _ensure_dir(año_dir)
        pedido_dir = año_dir / spec.folder_name
        _ensure_dir(pedido_dir)
        logger.info("Pedido creado: %s", pedido_dir)

    # 3) Subcarpeta técnica
    tecnico_dir = find_tecnico_dir(pedido_dir)
    if tecnico_dir is None:
        tecnico_dir = pedido_dir / "2-Tecnico"
        _ensure_dir(tecnico_dir)
        logger.info("2-Tecnico no existía — creado: %s", tecnico_dir)

    # 4) 00 DOCUMENTACIÓN
    doc_dir = tecnico_dir / "00 DOCUMENTACIÓN"
    _ensure_dir(doc_dir)

    return pedido_dir, tecnico_dir, doc_dir


def create_subfolders(
    tecnico_dir: Path,
    folder_names: list[str] | None = None,
) -> list[Path]:
    """Crea las subcarpetas `env. *` seleccionadas dentro de `2-Tecnico`.

    Si `folder_names` es None se crean todas las del catálogo.
    Las carpetas que ya existen se respetan (no se borran ni se vuelven a crear).
    Devuelve la lista de rutas (todas las solicitadas, existieran o no).
    """
    if not tecnico_dir.exists():
        raise FileNotFoundError(f"2-Tecnico no existe: {tecnico_dir}")

    names = folder_names if folder_names is not None else ALL_SUBFOLDERS
    created: list[Path] = []
    for name in names:
        if not name:
            continue
        target = tecnico_dir / name
        target.mkdir(parents=True, exist_ok=True)
        created.append(target)
    return created


def copy_documentation_template(
    documentacion_dir: Path,
    template_dir: Path = DEFAULT_TEMPLATE_DIR,
    overwrite: bool = False,
) -> list[str]:
    """Copia el contenido íntegro de la plantilla 00 DOCUMENTACIÓN al destino.

    Si `overwrite=False`, se respetan los archivos/carpetas que ya existen.
    Devuelve la lista de elementos top-level copiados.
    """
    if not template_dir.exists():
        raise FileNotFoundError(
            f"Plantilla 00 DOCUMENTACIÓN no encontrada en {template_dir}"
        )

    copied: list[str] = []
    for item in template_dir.iterdir():
        dst = documentacion_dir / item.name
        try:
            if item.is_dir():
                if dst.exists() and not overwrite:
                    # merge: copy children que no existan
                    for child in item.rglob("*"):
                        rel = child.relative_to(item)
                        target = dst / rel
                        if child.is_dir():
                            _ensure_dir(target)
                        elif not target.exists():
                            _ensure_dir(target.parent)
                            shutil.copy2(child, target)
                else:
                    if dst.exists() and overwrite:
                        shutil.rmtree(dst)
                    shutil.copytree(item, dst)
            else:
                if dst.exists() and not overwrite:
                    continue
                shutil.copy2(item, dst)
            copied.append(item.name)
        except Exception as exc:
            logger.exception("Fallo copiando %s → %s: %s", item, dst, exc)
            raise
    return copied


# ════════════════════════════════════════════════════════════════════════════
# Planning — copia el .xlsm y rellena la hoja PLAN ING
# ════════════════════════════════════════════════════════════════════════════


# Filas de fases en PLAN ING (template original):
# 7  PO sup0 date
# 8  eGesDoc opening
# 9  Critical Documentation
# 10 Non critical Documentation
# 11 Raw material and TT purchase
# 12 Mechanisation
# 13 Assembly and Testing
# 14 Delivery of TT
# 15 Final Assembly and Testing with TT
# 16 Inspection
# 17 Shipment of Material to Plant
_PHASE_ROWS = list(range(7, 18))


def _scale_phase_dates(
    template_start: datetime,
    template_end: datetime,
    new_start: datetime,
    new_end: datetime,
    template_phase: datetime,
) -> datetime:
    """Escala proporcionalmente una fecha de fase de plantilla al nuevo rango."""
    tpl_span = (template_end - template_start).total_seconds()
    new_span = (new_end - new_start).total_seconds()
    if tpl_span <= 0:
        return new_start
    offset = (template_phase - template_start).total_seconds()
    ratio = offset / tpl_span
    new_offset = ratio * new_span
    return new_start + timedelta(seconds=new_offset)


def generate_planning(
    spec: OrderSpec,
    documentacion_dir: Path,
    planning_template: Path = DEFAULT_PLANNING_TEMPLATE,
) -> Path:
    """Copia la plantilla Planning al pedido y rellena la hoja PLAN ING.

    Devuelve la ruta del Planning generado.
    """
    if not planning_template.exists():
        raise FileNotFoundError(
            f"Plantilla Planning no encontrada en {planning_template}"
        )

    dst_name = f"03 PLANNING - {spec.pedido}.xlsm"
    dst = documentacion_dir / dst_name

    shutil.copy2(planning_template, dst)

    # Abrir manteniendo macros (.xlsm)
    wb = load_workbook(dst, keep_vba=True, data_only=False)
    if "PLAN ING" not in wb.sheetnames:
        raise RuntimeError(
            f"La plantilla no contiene hoja 'PLAN ING'. Hojas: {wb.sheetnames}"
        )

    ws = wb["PLAN ING"]

    # Leer fechas y fases originales de la plantilla
    tpl_start = ws["G1"].value
    tpl_end = ws["G2"].value
    if not isinstance(tpl_start, datetime) or not isinstance(tpl_end, datetime):
        # Si la plantilla viene sin fechas, usar las nuevas tal cual sin escalar
        tpl_start = spec.fecha_entrada
        tpl_end = spec.fecha_prevista

    # Cabecera: fechas del rango global + S.REF + N.REF
    ws["G1"] = spec.fecha_entrada
    ws["G2"] = spec.fecha_prevista
    if spec.sref:
        ws["B2"] = spec.sref
    ws["B3"] = spec.n_ref_short

    # Escalar fases proporcionalmente
    for row in _PHASE_ROWS:
        b_cell = ws.cell(row=row, column=2)  # Fecha Inicio
        d_cell = ws.cell(row=row, column=4)  # Fecha Fin
        if isinstance(b_cell.value, datetime):
            b_cell.value = _scale_phase_dates(
                tpl_start, tpl_end, spec.fecha_entrada, spec.fecha_prevista, b_cell.value
            )
        if isinstance(d_cell.value, datetime):
            d_cell.value = _scale_phase_dates(
                tpl_start, tpl_end, spec.fecha_entrada, spec.fecha_prevista, d_cell.value
            )

    # GRAF ING — actualizar también las celdas de referencia visible
    if "GRAF ING" in wb.sheetnames:
        gws = wb["GRAF ING"]
        if spec.sref:
            gws["F4"] = spec.sref
        gws["F5"] = spec.n_ref  # con sufijo S00

    wb.save(dst)
    return dst


# ════════════════════════════════════════════════════════════════════════════
# VDDL — generado a partir de la plantilla ERP IMPORTAR
# ════════════════════════════════════════════════════════════════════════════
#
# Plantilla: M:\…\01 IMPORTAR DOCUMENTOS - ERP - sin enviar.xlsx
# Columnas: A=Nº Documento EIPSA · B=Nº Documento Cliente · C=Nº Pedido
#           D=Título · E=Tipo Documento · F=Crítico
#
# Cómo lo llenamos automáticamente:
#   - A, B, F → en blanco (el Document Controller los completa más tarde)
#   - C       → spec.n_ref ("P-XX/XXX-Sxx")
#   - D       → Título del catálogo según idioma seleccionado
#   - E       → "code" del catálogo (folder name sin "env. ")
#
# Una fila por subcarpeta tickada por el usuario, en el orden del catálogo.


def generate_vddl(
    spec: OrderSpec,
    documentacion_dir: Path,
    selected_folders: list[str] | None = None,
    lang: str = "es",
    erp_template: Path = DEFAULT_ERP_TEMPLATE,
) -> tuple[Path, list[str]]:
    """Genera el VDDL del pedido a partir de la plantilla ERP IMPORTAR.

    - `selected_folders`: nombres de las env. * tickadas. None = todas.
    - `lang`: 'es' o 'en' — controla la columna Título.
    - `erp_template`: ruta a la plantilla de importación ERP.

    Rellena:
      A · Nº Documento EIPSA = `{YY-NNN}-{eipsa_code}-{seq:04d}` (vacío si eipsa_code=None)
      C · Nº Pedido          = spec.n_ref ("P-XX/XXX-Sxx")
      D · Título             = ES o EN del catálogo
      E · Tipo Documento     = entrada del desplegable ERP (LITERAL)
      B, F                   = vacíos (Doc Controller los completa)

    Devuelve (ruta_destino, lista_de_warnings).
    """
    if not erp_template.exists():
        raise FileNotFoundError(
            f"Plantilla ERP no encontrada: {erp_template}"
        )

    if selected_folders is None:
        selected_folders = ALL_SUBFOLDERS

    # Conserva el orden del catálogo (no el de selección)
    chosen_entries: list[dict] = [
        SUBFOLDER_INDEX[f] for f in ALL_SUBFOLDERS
        if f in selected_folders and f in SUBFOLDER_INDEX
    ]

    dst = documentacion_dir / f"01 IMPORTAR DOCUMENTOS - ERP - {spec.pedido}.xlsx"
    shutil.copy2(erp_template, dst)

    wb = load_workbook(dst)
    ws = wb.active  # Hoja1

    # Borrar filas de ejemplo (mantener cabecera fila 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    n_ped = spec.n_ref          # "P-26/050-S00"
    prefix = spec.eipsa_prefix  # "26-050"

    missing_codes: list[str] = []
    invalid_tipos: list[str] = []
    invalid_criticos: list[str] = []
    missing_criticos: list[str] = []
    valid_dropdown = set(ERP_TIPO_DOC_OPTIONS)
    valid_critico = set(ERP_CRITICO_OPTIONS)

    for i, entry in enumerate(chosen_entries, start=2):
        # Nº Documento EIPSA = "{YY-NNN}-{eipsa_code}" → "26-050-LIS-0001"
        code = entry.get("eipsa_code")
        eipsa_num = f"{prefix}-{code}" if code else None
        if not code:
            missing_codes.append(entry["folder"])

        tipo = entry.get("tipo_doc") or ""
        if tipo and tipo not in valid_dropdown:
            invalid_tipos.append(f"{entry['folder']}={tipo!r}")

        critico = entry.get("critico") or ""
        if not critico:
            missing_criticos.append(entry["folder"])
        elif critico not in valid_critico:
            invalid_criticos.append(f"{entry['folder']}={critico!r}")

        ws.cell(row=i, column=1, value=eipsa_num)                     # Nº Doc EIPSA
        ws.cell(row=i, column=2, value=None)                          # Nº Doc Cliente
        ws.cell(row=i, column=3, value=n_ped)                         # Nº Pedido
        ws.cell(row=i, column=4, value=subfolder_title(entry, lang))  # Título
        ws.cell(row=i, column=5, value=tipo)                          # Tipo Documento (DROPDOWN)
        ws.cell(row=i, column=6, value=critico or None)               # Crítico (DROPDOWN Sí/No)

    wb.save(dst)

    warnings: list[str] = []
    if missing_codes:
        warnings.append(
            f"{len(missing_codes)} código(s) EIPSA pendientes (columna A vacía): "
            + ", ".join(missing_codes)
        )
    if invalid_tipos:
        warnings.append(
            f"{len(invalid_tipos)} tipo(s) Documento no coinciden con el desplegable: "
            + ", ".join(invalid_tipos)
        )
    if missing_criticos:
        warnings.append(
            f"{len(missing_criticos)} fila(s) sin Crítico (columna F vacía): "
            + ", ".join(missing_criticos)
        )
    if invalid_criticos:
        warnings.append(
            f"{len(invalid_criticos)} Crítico(s) no son 'Sí'/'No': "
            + ", ".join(invalid_criticos)
        )
    return dst, warnings


# ════════════════════════════════════════════════════════════════════════════
# Orquestador — punto de entrada principal
# ════════════════════════════════════════════════════════════════════════════


def create_order(
    pedido: str,
    fecha_entrada: datetime,
    fecha_prevista: datetime,
    *,
    cliente: str = "",
    material: str = "",
    sref: str = "",
    año: int | None = None,
    suffix: str = "S00",
    base_dir: Path | str = DEFAULT_BASE_DIR,
    template_dir: Path | str = DEFAULT_TEMPLATE_DIR,
    planning_template: Path | str = DEFAULT_PLANNING_TEMPLATE,
    erp_template: Path | str = DEFAULT_ERP_TEMPLATE,
    copy_template: bool = True,
    do_planning: bool = True,
    do_vddl: bool = True,
    do_subfolders: bool = True,
    overwrite_template: bool = False,
    create_if_missing: bool = False,
    subfolders: list[str] | None = None,
    vddl_lang: str = "es",
) -> OrderResult:
    """Localiza el pedido y añade 00 DOCUMENTACIÓN + Planning + VDDL.

    Por defecto **no crea** la carpeta del pedido si no existe (sólo añade
    documentación a un pedido ya creado). Para crear la carpeta completa pasa
    `create_if_missing=True` (requiere cliente y material).
    """
    spec = OrderSpec(
        pedido=pedido,
        cliente=cliente,
        material=material,
        fecha_entrada=fecha_entrada,
        fecha_prevista=fecha_prevista,
        sref=sref,
        año=año,
        suffix=suffix,
    )

    base_dir = Path(base_dir)
    template_dir = Path(template_dir)
    planning_template = Path(planning_template)
    erp_template = Path(erp_template)
    lang = (vddl_lang or "es").lower()
    if lang not in ("es", "en"):
        lang = "es"

    warnings: list[str] = []

    # 1) Localizar (o crear) carpetas
    pedido_dir, tecnico_dir, doc_dir = create_folder_structure(
        spec, base_dir=base_dir, create_if_missing=create_if_missing,
    )
    logger.info("00 DOCUMENTACIÓN preparado en: %s", doc_dir)

    result = OrderResult(
        pedido_dir=pedido_dir,
        tecnico_dir=tecnico_dir,
        documentacion_dir=doc_dir,
    )

    # 2) Copia plantilla 00 DOCUMENTACIÓN
    if copy_template:
        try:
            copied = copy_documentation_template(
                doc_dir, template_dir=template_dir, overwrite=overwrite_template
            )
            result.template_copied = True
            logger.info("Plantilla copiada: %d elementos top-level.", len(copied))
        except FileNotFoundError as exc:
            warnings.append(f"No se pudo copiar plantilla: {exc}")
            logger.warning("Plantilla no copiada: %s", exc)

    # 3) Planning
    if do_planning:
        try:
            tpl = planning_template
            # Si copiamos plantilla, preferimos la copia recién creada en doc_dir
            local_tpl = doc_dir / "03 PLANTILLA PLANNING - CON MESES.xlsm"
            if local_tpl.exists():
                tpl = local_tpl
            result.planning_file = generate_planning(
                spec, doc_dir, planning_template=tpl
            )
            logger.info("Planning generado: %s", result.planning_file)
        except Exception as exc:
            warnings.append(f"No se pudo generar Planning: {exc}")
            logger.exception("Fallo generando Planning")

    # 4) Subcarpetas env. * dentro de 2-Tecnico
    if do_subfolders:
        try:
            created = create_subfolders(tecnico_dir, subfolders)
            result.subfolders_created = [p.name for p in created]
            logger.info("Subcarpetas creadas/garantizadas: %d", len(created))
        except Exception as exc:
            warnings.append(f"No se pudieron crear subcarpetas: {exc}")
            logger.exception("Fallo creando subcarpetas")

    # 5) VDDL — generado desde plantilla ERP, una fila por subcarpeta tickada
    if do_vddl:
        try:
            vddl_path, vddl_warnings = generate_vddl(
                spec, doc_dir,
                selected_folders=subfolders,
                lang=lang,
                erp_template=erp_template,
            )
            result.vddl_file = vddl_path
            result.vddl_lang = lang
            warnings.extend(vddl_warnings)
            logger.info("VDDL generado: %s (lang=%s)", vddl_path, lang)
        except Exception as exc:
            warnings.append(f"No se pudo generar VDDL: {exc}")
            logger.exception("Fallo generando VDDL")

    result.warnings = warnings
    return result
